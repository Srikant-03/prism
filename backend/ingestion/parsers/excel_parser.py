"""
Excel parser supporting .xlsx, .xls, .xlsm formats.
Multi-sheet discovery, merged cell handling, and sheet selection.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any, Callable, Optional

import pandas as pd
import openpyxl

from models.schemas import MalformedReport, MalformedRow, MalformedSeverity, SheetInfo
from ingestion.parsers.base import BaseParser, ParseResult


class ExcelParser(BaseParser):
    """
    Parser for Excel files (.xlsx, .xls, .xlsm).
    Discovers all sheets, handles merged cells, and supports sheet selection.
    """

    def __init__(self, file_path: Path, encoding: str = "utf-8"):
        super().__init__(file_path, encoding)
        self._sheets_info: Optional[list[SheetInfo]] = None

    def validate(self) -> tuple[bool, str]:
        """Check that the file is a valid Excel file."""
        if not self.file_path.exists():
            return False, f"File not found: {self.file_path}"
        if self.file_path.stat().st_size == 0:
            return False, "File is empty"

        ext = self.file_path.suffix.lower()
        if ext == ".xls":
            try:
                import xlrd
                xlrd.open_workbook(str(self.file_path))
                return True, "Valid .xls file"
            except Exception as e:
                return False, f"Invalid .xls file: {str(e)}"
        else:
            try:
                wb = openpyxl.load_workbook(str(self.file_path), read_only=True)
                wb.close()
                return True, f"Valid {ext} file"
            except Exception as e:
                return False, f"Invalid Excel file: {str(e)}"

    def get_metadata(self) -> dict:
        """Return sheet information."""
        sheets = self.discover_sheets()
        return {
            "sheet_count": len(sheets),
            "sheets": [s.model_dump() for s in sheets],
        }

    def discover_sheets(self) -> list[SheetInfo]:
        """
        Discover all sheets in the workbook with metadata:
        name, row/col count, merged cell presence, and preview rows.
        """
        if self._sheets_info is not None:
            return self._sheets_info

        ext = self.file_path.suffix.lower()
        if ext == ".xls":
            return self._discover_sheets_xls()
        return self._discover_sheets_openpyxl()

    def _discover_sheets_openpyxl(self) -> list[SheetInfo]:
        """Discover sheets using openpyxl (for .xlsx, .xlsm)."""
        sheets = []
        try:
            wb = openpyxl.load_workbook(str(self.file_path), read_only=False, data_only=True)

            for idx, name in enumerate(wb.sheetnames):
                ws = wb[name]
                has_merged = len(ws.merged_cells.ranges) > 0

                # Get preview rows (first 5)
                preview = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 5:
                        break
                    preview.append([self._cell_value(v) for v in row])

                sheets.append(SheetInfo(
                    name=name,
                    index=idx,
                    row_count=ws.max_row or 0,
                    col_count=ws.max_column or 0,
                    has_merged_cells=has_merged,
                    preview=preview,
                ))

            wb.close()
        except Exception as e:
            sheets.append(SheetInfo(
                name="Error",
                index=0,
                row_count=0,
                col_count=0,
                preview=[[f"Error reading sheets: {str(e)}"]],
            ))

        self._sheets_info = sheets
        return sheets

    def _discover_sheets_xls(self) -> list[SheetInfo]:
        """Discover sheets using xlrd (for .xls)."""
        import xlrd
        sheets = []
        try:
            wb = xlrd.open_workbook(str(self.file_path))
            for idx in range(wb.nsheets):
                ws = wb.sheet_by_index(idx)
                preview = []
                for i in range(min(5, ws.nrows)):
                    preview.append([self._cell_value(ws.cell_value(i, j)) for j in range(ws.ncols)])

                sheets.append(SheetInfo(
                    name=ws.name,
                    index=idx,
                    row_count=ws.nrows,
                    col_count=ws.ncols,
                    has_merged_cells=len(ws.merged_cells) > 0 if hasattr(ws, "merged_cells") else False,
                    preview=preview,
                ))
        except Exception as e:
            sheets.append(SheetInfo(
                name="Error",
                index=0,
                row_count=0,
                col_count=0,
                preview=[[f"Error reading sheets: {str(e)}"]],
            ))

        self._sheets_info = sheets
        return sheets

    def parse(self, selected_sheets: Optional[list[int]] = None, **kwargs) -> ParseResult:
        """
        Parse the Excel file. If selected_sheets is None and there are multiple sheets,
        returns metadata for sheet selection. Otherwise, parses selected sheet(s).
        """
        sheets = self.discover_sheets()

        # If multiple sheets and none selected, prompt user for selection
        if len(sheets) > 1 and selected_sheets is None:
            return ParseResult(
                metadata={
                    **self.get_metadata(),
                    "requires_sheet_selection": True,
                },
                justification=(
                    f"Excel file contains {len(sheets)} sheets: "
                    f"{', '.join(s.name for s in sheets)}. "
                    f"Please select which sheet(s) to analyze."
                ),
            )

        # Default to first sheet if single sheet
        if selected_sheets is None:
            selected_sheets = [0]

        justification_parts = []
        all_dfs = {}
        malformed_issues = []

        ext = self.file_path.suffix.lower()

        for sheet_idx in selected_sheets:
            if sheet_idx >= len(sheets):
                malformed_issues.append(MalformedRow(
                    row_number=0,
                    raw_content="",
                    issue=f"Sheet index {sheet_idx} does not exist (only {len(sheets)} sheets available)",
                    severity=MalformedSeverity.ERROR,
                ))
                continue

            sheet_name = sheets[sheet_idx].name
            try:
                # Handle merged cells for xlsx/xlsm
                if ext != ".xls" and sheets[sheet_idx].has_merged_cells:
                    df = self._parse_with_merged_cells(sheet_name)
                    justification_parts.append(
                        f"Sheet '{sheet_name}': Unmerged and filled {sheets[sheet_idx].row_count} "
                        f"rows × {sheets[sheet_idx].col_count} columns with merged cell handling."
                    )
                else:
                    engine = "xlrd" if ext == ".xls" else "openpyxl"
                    df = pd.read_excel(
                        str(self.file_path),
                        sheet_name=sheet_name,
                        engine=engine,
                    )
                    justification_parts.append(
                        f"Sheet '{sheet_name}': Parsed {len(df)} rows × {len(df.columns)} columns."
                    )

                all_dfs[sheet_name] = df

            except Exception as e:
                malformed_issues.append(MalformedRow(
                    row_number=0,
                    raw_content="",
                    issue=f"Failed to parse sheet '{sheet_name}': {str(e)}",
                    severity=MalformedSeverity.ERROR,
                ))

        # Combine if multiple sheets selected
        if len(all_dfs) == 1:
            result_df = list(all_dfs.values())[0]
        elif len(all_dfs) > 1:
            # Stack sheets with a source column
            combined = []
            for name, df in all_dfs.items():
                df = df.copy()
                df["__source_sheet__"] = name
                combined.append(df)
            result_df = pd.concat(combined, ignore_index=True)
            justification_parts.append(
                f"Combined {len(all_dfs)} sheets into a single DataFrame with "
                f"'__source_sheet__' column for traceability."
            )
        else:
            result_df = pd.DataFrame()

        malformed = MalformedReport(
            has_issues=len(malformed_issues) > 0,
            total_issues=len(malformed_issues),
            issues=malformed_issues,
            summary=f"{len(malformed_issues)} sheet parsing issues" if malformed_issues else "",
        )

        return ParseResult(
            dataframe=result_df,
            metadata=self.get_metadata(),
            malformed=malformed,
            justification=" ".join(justification_parts),
        )

    def _parse_with_merged_cells(self, sheet_name: str) -> pd.DataFrame:
        """
        Parse an Excel sheet while properly handling merged cells.
        Merged cells are unmerged and filled with the merged value.
        """
        wb = openpyxl.load_workbook(str(self.file_path), data_only=True)
        ws = wb[sheet_name]

        # Record merged cell ranges before unmerging
        merged_ranges = list(ws.merged_cells.ranges)

        # Unmerge all cells
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))

            # Fill all cells in the range with the top-left value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col, value=top_left_value)

        # Convert to DataFrame
        data = list(ws.values)
        wb.close()

        if not data:
            return pd.DataFrame()

        # First row as header
        header = [self._cell_value(v) for v in data[0]]
        # Ensure unique column names
        seen = {}
        unique_header = []
        for h in header:
            h_str = str(h) if h is not None else "Unnamed"
            if h_str in seen:
                seen[h_str] += 1
                unique_header.append(f"{h_str}_{seen[h_str]}")
            else:
                seen[h_str] = 0
                unique_header.append(h_str)

        rows = [[self._cell_value(v) for v in row] for row in data[1:]]
        return pd.DataFrame(rows, columns=unique_header)

    @staticmethod
    def _cell_value(value: Any) -> Any:
        """Convert cell value to a JSON-safe type."""
        if value is None:
            return None
        if isinstance(value, (int, float, str, bool)):
            return value
        return str(value)
